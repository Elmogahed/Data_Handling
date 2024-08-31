from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

data = [
    ['Samir', 98, 95, 90, 93],
    ['Ahmed', 98, 95, 90, 93],
    ['Malek', 98, 95, 90, 93],
    ['Mansour', 98, 95, 90, 93],
]



wb = Workbook()
ws = wb.active


for row in data: 
    ws.append(row)



for i in range(1, len(data)+1):
    ws.cell(row=i, column=6, value=f'=AVERAGE(B{i}:E{i})')


values = Reference(ws, min_col=1, min_row=1, max_col=5, max_row=4)
names = Reference(ws, min_col=1, min_row=1, max_col=1, max_row=4)

chart = BarChart()
chart.type = 'bar'
chart.title = 'Students Degrees'
chart.y_axis.title = 'Degree'
chart.x_axis.title = 'Student'
chart.width = 20
chart.height = 15
chart.legend = None


chart.add_data(values)
ws.add_chart(chart, 'E15')

ws1 = wb.create_sheet('MySheet')
ws2 = wb.create_sheet('MySheet', 0)
ws3 = wb.create_sheet('MySheet', -1)
ws = wb['MySheet']

cell = ws['A4']

ws['A4'] = 11

ws.cell(row=4, column=2, value=11)

wb.save('excel.xlsx')
